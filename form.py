# -*- coding: utf-8 -*-
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.support.ui import Select

from time import sleep

import openpyxl
import os
import validaciones as v


driver = webdriver.Chrome(executable_path='./chromedriver')
wait = WebDriverWait(driver, 10)

url ='https://docs.google.com/forms/d/e/1FAIpQLSfGVV8gfqTaLaQzWq03n4bBsK0UJ_xCuMmiwOFYBbpJnqNjLQ/formResponse'

def record_data_form(data, url):
  time = 0.5
  time1 = 0.3
  driver.get(url)
  sleep(time)

  # PRIMERA SECCIÓN
  # 1. Número contrato
  driver.find_element_by_css_selector("input[aria-label='1. Número de contrato']").send_keys('4148.010.26.1131-2017')
  # 2. Código proyecto
  driver.find_element_by_css_selector("input[aria-label='2. Número o código del Proyecto']").send_keys('06046296')
  # 3. Fecha
  el = driver.find_element_by_xpath("//input[@type='date']")
  el.click()
  el.send_keys(data[5])
  # 4. Actividad que se realiza
  driver.find_element_by_css_selector("input[aria-label='4. Nombre de la actividad que se realiza']").send_keys(data[6])
  # 5. Lugar de la actividad
  driver.find_element_by_css_selector("input[aria-label='5. Lugar donde se desarrolla la actividad']").send_keys(data[2])
  # 6. Responsable
  driver.find_element_by_css_selector("input[aria-label='6. Persona responsable de diligenciar el registro de actividades']").send_keys(data[7])
  sleep(time) 
  # 7. Tipo Beneficio
  driver.find_element_by_css_selector("div[role=listbox]").click()
  tipo_beneficio_dropdown = ".exportSelectPopup .quantumWizMenuPaperselectOption[data-value='{}']".format(data[8])
  wait.until(EC.visibility_of_element_located((By.CSS_SELECTOR, tipo_beneficio_dropdown)))
  driver.find_element_by_css_selector(tipo_beneficio_dropdown).click()
  # Boton Siguiente
  sleep(time)
  driver.find_element_by_xpath("//span[text()='Siguiente']").click()
  
  # SEGUNDA SECCIÓN
  # 8. Tipo Documento de Identidad
  driver.find_element_by_xpath("//span[.='Cédula de ciudadanía (CC)']").click()
  sleep(time1)
  # 9 Numero Documento
  driver.find_element_by_css_selector("input[aria-label='9. Nº de documento']").send_keys(data[10])
  # 10. Primer Nombre
  if not data[11] == None:
    driver.find_element_by_css_selector("input[aria-label='10. Primer Nombre']").send_keys(data[11].capitalize())
  # 11. Segundo Nombre
  if not data[12] == None:
    driver.find_element_by_css_selector("input[aria-label='11. Segundo Nombre']").send_keys(data[12].capitalize())
  # 12. Primer Apellido
  if not data[13] == None:
    driver.find_element_by_css_selector("input[aria-label='12. Primer Apellido']").send_keys(data[13].capitalize())
  # 13. Segundo Apellido
  if not data[14] == None:
    driver.find_element_by_css_selector("input[aria-label='13. Segundo Apellido']").send_keys(data[14].capitalize())
  # 14. Edad
  driver.find_element_by_css_selector("input[aria-label='14. Edad']").send_keys(data[15])
  # 16. Sexo
  if int(data[16]) == 1:
    driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[2]/form[1]/div[1]/div[2]/div[2]/div[9]/div[1]/div[2]/div[1]/span[1]/div[1]/div[1]/label[1]/div[1]/div[2]/div[1]/span[1]").click()
  elif int(data[16]) == 2:
    driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[2]/form[1]/div[1]/div[2]/div[2]/div[9]/div[1]/div[2]/div[1]/span[1]/div[1]/div[2]/label[1]/div[1]/div[2]/div[1]/span[1]").click()
  else:
    print('error sexo')
  # 17. LGTBI
  if int(data[17]) == 1:
    driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[2]/form[1]/div[1]/div[2]/div[2]/div[10]/div[1]/div[2]/div[1]/span[1]/div[1]/div[1]/label[1]/div[1]/div[2]/div[1]/span[1]").click()
  elif int(data[17]) == 2:
    driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[2]/form[1]/div[1]/div[2]/div[2]/div[10]/div[1]/div[2]/div[1]/span[1]/div[1]/div[2]/label[1]/div[1]/div[2]/div[1]/span[1]").click()
  else:
    print('error lgtbi')


  # 18. Grupo Etnico al cual pertenece
  if int(data[18]) == 1:
    driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[2]/form[1]/div[1]/div[2]/div[2]/div[11]/div[1]/div[2]/div[1]/span[1]/div[1]/div[1]/label[1]/div[1]/div[2]/div[1]/span[1]").click()
  elif int(data[18]) == 2:
    driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[2]/form[1]/div[1]/div[2]/div[2]/div[11]/div[1]/div[2]/div[1]/span[1]/div[1]/div[2]/label[1]/div[1]/div[2]/div[1]/span[1]").click()
  elif int(data[18]) == 3:
    driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[2]/form[1]/div[1]/div[2]/div[2]/div[11]/div[1]/div[2]/div[1]/span[1]/div[1]/div[3]/label[1]/div[1]/div[2]/div[1]/span[1]").click()
  elif int(data[18]) == 4:
    driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[2]/form[1]/div[1]/div[2]/div[2]/div[11]/div[1]/div[2]/div[1]/span[1]/div[1]/div[4]/label[1]/div[1]/div[2]/div[1]/span[1]").click()
  elif int(data[18]) == 5:
    driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[2]/form[1]/div[1]/div[2]/div[2]/div[11]/div[1]/div[2]/div[1]/span[1]/div[1]/div[5]/label[1]/div[1]/div[2]/div[1]/span[1]").click()
  elif int(data[18]) == 6:
    driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[2]/form[1]/div[1]/div[2]/div[2]/div[11]/div[1]/div[2]/div[1]/span[1]/div[1]/div[6]/label[1]/div[1]/div[2]/div[1]/span[1]").click()
  elif int(data[18]) == 7:
    driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[2]/form[1]/div[1]/div[2]/div[2]/div[11]/div[1]/div[2]/div[1]/span[1]/div[1]/div[7]/label[1]/div[1]/div[2]/div[1]/span[1]").click()
  else:
    driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[2]/form[1]/div[1]/div[2]/div[2]/div[11]/div[1]/div[2]/div[1]/span[1]/div[1]/div[8]/label[1]/div[1]/div[2]/div[1]/span[1]").click()

  # Boton Siguiente
  sleep(time)
  driver.find_element_by_xpath("//span[text()='Siguiente']").click()

  # SECCIÓN TRES
  # Tipo de Discapacidad
  if data[19] == None:
    # Boton Siguiente
    sleep(time)
    driver.find_element_by_xpath("//span[text()='Siguiente']").click()
  if data[19] != None:
    tipo_discapacidad = v.convertir_tipo_discapacidad(data[19])
    driver.find_element_by_css_selector("div[aria-label='{}']".format(tipo_discapacidad)).click()
    driver.find_element_by_xpath("//span[text()='Siguiente']").click()

  

  # SECCIÓN CUATRO
  # Victima de conflicto
  conflicto = v.convertir_si_no(int(data[20]))
  driver.find_element_by_css_selector("div[aria-label='{}']".format(conflicto)).click()
  # Boton Siguiente
  sleep(time)
  driver.find_element_by_xpath("//span[text()='Siguiente']").click()

  # SECCIÓN CINCO
  # Comuna
  driver.find_element_by_css_selector("div[role=listbox]").click()
  comuna_dropdown = ".exportSelectPopup .quantumWizMenuPaperselectOption[data-value='{}']".format(data[22])
  wait.until(EC.visibility_of_element_located((By.CSS_SELECTOR, comuna_dropdown)))
  driver.find_element_by_css_selector(comuna_dropdown).click()
  sleep(time)
  # Boton Siguiente
  driver.find_element_by_xpath("//span[text()='Siguiente']").click()

  # SECCIÓN SEIS
  # Sector
  driver.find_element_by_css_selector("div[role=listbox]").click() 
  sector_dropdown = ".exportSelectPopup .quantumWizMenuPaperselectOption[data-value='{}']".format(data[23])
  wait.until(EC.visibility_of_element_located((By.CSS_SELECTOR, sector_dropdown)))
  driver.find_element_by_css_selector(sector_dropdown).click()
  # Boton Siguiente
  sleep(time)
  driver.find_element_by_xpath("//span[text()='Siguiente']").click()

  # SECCIÓN SIETE
  # Telefono
  if not (data[21] == None or data[21] == " "):
    driver.find_element_by_css_selector("input[aria-label='24. Número de teléfono fijo o celular']").send_keys(int(data[21]))
  # Correo
  if not (data[24] == None or data[24] == " "):
    driver.find_element_by_css_selector("input[aria-label='25. Correo electrónico']").send_keys(data[24])

  # FIN FORMULARIO
  sleep(time1)
  driver.find_element_by_xpath("//span[text()='Enviar']").click()

# Read file excel
path = './BD1.xlsx'
SHEET_NAME = 'TOTAL BENEFICIARIOS EN IEO'

# Cargar Archivo
workbook = openpyxl.load_workbook(path)
sheet = workbook.get_sheet_by_name(SHEET_NAME)

# Obtener filas del archivo
rows = sheet.max_row

def leer_archivo_excel(url, workbook, sheet, cols, fila_inicial, fila_final):

  for r in range(fila_inicial, fila_final+1):
    print(r)
    data = ['']
    for c in range(1, cols+1):
      data.append(sheet.cell(row=r, column=c).value)
      
    record_data_form(data, url)
    # print(data)
    print("data success")
  print('end')

leer_archivo_excel(url, workbook, sheet, 29, 801, 900)
