# -*- coding: utf-8 -*-
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.support.ui import Select

from time import sleep

import openpyxl
import os
import validaciones as v


driver = webdriver.Chrome(executable_path='./chromedriver')
# driver = webdriver.Firefox(executable_path='./chromedriver')
wait = WebDriverWait(driver, 15)

url = 'https://docs.google.com/forms/d/e/1FAIpQLSd_oXWV8r9pvgmdTJR8ARjblvXGbOhqf4-AX241dinzyR_Bwg/viewform?fbzx=-673584793312408014'

# Record data to form
def record_data_form(data, url):

  time = .8
  # Cargar página
  driver.get(url)
  sleep(3)


  # PRIMERA SECCIÓN

  # Boton Siguiente
  driver.find_element_by_xpath("//span[text()='Siguiente']").click()


  # SEGUNDA SECCIÓN

  # Codigo proyecto
  driver.find_element_by_css_selector("input[aria-label='1. Número o código del proyecto (BP)']").send_keys('06046293')
  # Numero contrato
  driver.find_element_by_css_selector("input[aria-label='2. Número de contrato']").send_keys('4148.010.26.1131-2017')
  sleep(1)
  # Area o Disciplina
  driver.find_element_by_css_selector("div[role=listbox]").click() 
  area_dropdown = ".exportSelectPopup .quantumWizMenuPaperselectOption[data-value='{}']".format(data[5])
  wait.until(EC.visibility_of_element_located((By.CSS_SELECTOR, area_dropdown)))
  driver.find_element_by_css_selector(area_dropdown).click()
  
  # Boton Siguiente
  sleep(time)
  driver.find_element_by_xpath("//span[text()='Siguiente']").click()
  

  # TERCERA SECCIÓN
  
  # Tipo de documento
  tipo_documento = v.convertir_tipo_documento(int(data[6]))
  driver.find_element_by_css_selector(".freebirdFormviewerViewItemsCheckboxChoice div[aria-label='{}']".format(tipo_documento)).click()

  sleep(time)
  
  # Numero documento  
  driver.find_element_by_css_selector("input[aria-label='5. Nº de documento']").send_keys(int(data[7]))
  # Primer Nombre
  if not data[8] == None:
    driver.find_element_by_css_selector("input[aria-label='6. Primer Nombre']").send_keys(data[8].capitalize())
  # Segundo Nombre
  if not data[9] == None:
    driver.find_element_by_css_selector("input[aria-label='7. Segundo Nombre']").send_keys(data[9].capitalize())
  # Primer Apellido
  if not data[10] == None:
    driver.find_element_by_css_selector("input[aria-label='8. Primer Apellido']").send_keys(data[10].capitalize())
  # Segundo Apellido
  if not data[11] == None:
    driver.find_element_by_css_selector("input[aria-label='9. Segundo Apellido']").send_keys(data[11].capitalize())
  # Fecha de nacimiento
  # generar fecha
  # fecha = v.generar_fecha(data[12])
  # print(fecha)
  el = driver.find_element_by_xpath("//input[@type='date']")
  el.click()
  # fecha = data[12].strftime("%m/%d/%y")
  sleep(.1)
  el.send_keys(data[12])
  sleep(.1)
  # Lugar de nacimiento
  driver.find_element_by_css_selector("input[aria-label='11. Lugar de nacimiento']").send_keys(data[13])
  # Sexo
  sexo = v.convertir_sexo(int(data[14]))
  driver.find_element_by_css_selector(".freebirdFormviewerViewItemsCheckboxChoice div[aria-label='{}']".format(sexo)).click()

  # Boton Siguiente
  driver.find_element_by_xpath("//span[text()='Siguiente']").click()
  sleep(.2)

  # CUARTA SECCIÓN

  # LGBTI
  lgbti = v.convertir_si_no(int(data[15]))
  sleep(.2)
  driver.find_element_by_css_selector("div[aria-label='{}']".format(lgbti)).click()
  sleep(.2)
  # Boton Siguiente
  driver.find_element_by_xpath("//span[text()='Siguiente']").click()


  # QUINTA SECCIÓN

  # Grupo Etnico
  sleep(time)
  etnia = v.convertir_etnia(data[16])
  driver.find_element_by_css_selector(".freebirdFormviewerViewItemsCheckboxChoice div[aria-label='{}']".format(etnia)).click()
  
  # Boton Siguiente
  driver.find_element_by_xpath("//span[text()='Siguiente']").click()


  # SEXTA SECCIÓN

  # Discapacidad
  sleep(.5)
  discapacidad = v.convertir_si_no(int(data[17]))
  driver.find_element_by_css_selector("div[aria-label='{}']".format(discapacidad)).click()
  # Boton Siguiente
  driver.find_element_by_xpath("//span[text()='Siguiente']").click()
  sleep(.2)

  # Si la respuesta es afirmativa, se debe completar otros datos, por lo tanto
  if discapacidad.lower() == 'si':
    # print(type(data[19]))
    tipo_discapacidad = v.convertir_tipo_discapacidad(data[18])
    driver.find_element_by_css_selector("div[aria-label='{}']".format(tipo_discapacidad)).click()    
    driver.find_element_by_xpath("//span[text()='Siguiente']").click()
    sleep(.2)


  # SEPTIMA SECCIÓN

  # Servicio de salud
  driver.find_element_by_css_selector("div[aria-label='No']").click()
  # Boton Siguiente
  driver.find_element_by_xpath("//span[text()='Siguiente']").click()
  sleep(.2)

  # # Si la respuesta es afirmativa, se deben completar otros datos, por lo tanto:
  # if data[15].lower() == 'si':
  #   # Tipo servicio de salud
  #   driver.find_element_by_css_selector("div[aria-label='{}']".format('Subsidiado')).click()
  #   sleep(2)
  #   # Boton Siguiente
  #   driver.find_element_by_xpath("//span[text()='Siguiente']").click()
    
  #   # Cual?
  #   driver.find_element_by_css_selector("textarea[aria-label='19. ¿Cuál?']").send_keys('Emsanar')
  #   sleep(2)
  #   # Boton Siguiente
  #   driver.find_element_by_xpath("//span[text()='Siguiente']").click()


  # OCTAVA SECCIÓN

  # Victima de conflicto
  conflicto = v.convertir_si_no(int(data[20]))
  driver.find_element_by_css_selector("div[aria-label='{}']".format(conflicto)).click()
  # Estudiante
  # Como hay dos grupos de radio button valido el tipo de dato para asi seleccionar
  # el segundo grupo de radio buttons correspondiente, ademas si la respuesta es SI
  # Se debe llenar otro campo
  driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[2]/form[1]/div[1]/div[2]/div[2]/div[3]/div[1]/div[2]/div[1]/span[1]/div[1]/div[1]/label[1]/div[1]/div[1]").click()
  # Boton Siguiente
  driver.find_element_by_xpath("//span[text()='Siguiente']").click()
  sleep(.2)

  # En caso afirmativo que Estudie INSTITUCIÓN EDUCATIVA
  # Institución educativa
  # Revisar el excel, PARA VERIFICAR QUE ESTE CAMPO EXISTE
  driver.find_element_by_css_selector("input[aria-label='22. Institución educativa']").send_keys(data[2])
  
  # Boton Siguiente
  sleep(time)
  driver.find_element_by_xpath("//span[text()='Siguiente']").click()
 
  

  # NOVENA SECCIÓN

  # Nivel educativo
  sleep(2)
  driver.find_element_by_css_selector("div[role=listbox]").click() 
  nivel_educativo_dropdown = ".exportSelectPopup .quantumWizMenuPaperselectOption[data-value='{}']".format('Básica Secundaria (Bachillerato básico)')
  wait.until(EC.visibility_of_element_located((By.CSS_SELECTOR, nivel_educativo_dropdown)))
  driver.find_element_by_css_selector(nivel_educativo_dropdown).click()
  
  sleep(time)
  # Boton Siguiente
  driver.find_element_by_xpath("//span[text()='Siguiente']").click()
  


  # DECIMA SECCIÓN

  # Comuna
  driver.find_element_by_css_selector("div[role=listbox]").click()
  if data[24] == 'No vive en Cali':
    driver.find_element_by_css_selector("input[aria-label='24.1- Nombre del municipio donde vive ']").send_keys()
    driver.find_element_by_css_selector("input[aria-label='24.2- Nombre del barrio donde vive ']").send_keys()

  comuna_dropdown = ".exportSelectPopup .quantumWizMenuPaperselectOption[data-value='{}']".format(data[24])
  wait.until(EC.visibility_of_element_located((By.CSS_SELECTOR, comuna_dropdown)))
  driver.find_element_by_css_selector(comuna_dropdown).click()
  sleep(time)
  # Boton Siguiente
  driver.find_element_by_xpath("//span[text()='Siguiente']").click()
  


  # ONCEAVA SECCIÓN

  # Barrio
  driver.find_element_by_css_selector("div[role=listbox]").click() 
  comuna_dropdown = ".exportSelectPopup .quantumWizMenuPaperselectOption[data-value='{}']".format(data[25])
  wait.until(EC.visibility_of_element_located((By.CSS_SELECTOR, comuna_dropdown)))
  driver.find_element_by_css_selector(comuna_dropdown).click()

  # Boton Siguiente
  sleep(time)
  driver.find_element_by_xpath("//span[text()='Siguiente']").click()
  

  # DOCEAVA SECCIÓN

  # Dirección
  # driver.find_element_by_css_selector("input[aria-label='26. Dirección']").send_keys(data[21])
  # Telefono
  if not data[26] == None:
    driver.find_element_by_css_selector("input[aria-label='27. Nº Teléfono fijo o celular']").send_keys(int(data[26]))
  # Correo
  if not data[27] == None:
    driver.find_element_by_css_selector("input[aria-label='28. Correo electrónico']").send_keys(data[27])
  # Nombre
  if not data[28] == None:
    driver.find_element_by_css_selector("input[aria-label='29. Nombre de la persona que elabora la inscripción']").send_keys(data[28])
  # Observacion
  # driver.find_element_by_css_selector("input[aria-label='30. Observación']").send_keys(data[25])

  # FIN FORMULARIO
  sleep(.2)
  driver.find_element_by_xpath("//span[text()='Enviar']").click()
  
# Read file excel
path = './CB1.xlsx'
SHEET_NAME = 'TOTAL BENEFICIARIOS EN IEO'

# Cargar Archivo
workbook = openpyxl.load_workbook(path)
sheet = workbook.get_sheet_by_name(SHEET_NAME)

# Obtener filas del archivo
rows = sheet.max_row

def leer_archivo_excel(url, workbook, sheet, cols, fila_inicial, fila_final):

  file = open('./log.txt', "w")
  
  for r in range(fila_inicial, fila_final+1):
    print(r)
    data = ['']
    for c in range(1, cols+1):
      data.append(sheet.cell(row=r, column=c).value)

    record_data_form(data, url)
    print("data success")
    sleep(0.5)
    
    
  
  print('end')

leer_archivo_excel(url, workbook, sheet, 29, 5505, 6000)
# Comenzar a partir del registro 4013
# Laura termina la fila 5000
# pino comienza 5001 hasta la 5500 - Terminado
# pino 5501 - 6000