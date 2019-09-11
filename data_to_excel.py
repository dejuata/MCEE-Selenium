# -*- coding: utf-8 -*-
import validaciones
import openpyxl
import os

from selenium import webdriver


# Read file excel
path = './BD1.xlsx'
# path = './CB1.xlsx'
SHEET_NAME = 'TOTAL BENEFICIARIOS EN IEO'

# Cargar Archivo
workbook = openpyxl.load_workbook(path)
sheet = workbook.get_sheet_by_name(SHEET_NAME)

# Obtener filas del archivo
rows = sheet.max_row
# Obtener columnas del archivo
# cols = sheet.max_column

def completar_comuna(workbook, sheet, column, rows):

  count = 0

  for r in range(2, rows+1):
    data_cell = sheet.cell(row=r, column=column).value
    
    if not data_cell == None:
      if data_cell.lower() == "V. BLANCA".lower():
        sheet.cell(row=r, column=column+1).value = "57. Pichinde"
        sheet.cell(row=r, column=column+2).value = "Peñas Blancas"

    comuna = sheet.cell(row=r, column=column+1).value
    if (comuna == None or comuna == " "):
      count += 1

  # Cerrar archivo de texto y excel
  workbook.save(path)
  print('Faltan: {}'.format(count))
  print("data success")

# completar_comuna(workbook, sheet, 23, 6308)

def validar_numero(workbook, sheet, column, rows, log):
  
  file = open(log, "w")

  for r in range(593, rows+1):
    print(r)
    data_cell = sheet.cell(row=r, column=column).value
    fecha = validaciones.generar_edad(data_cell)
    if fecha < 21:
      fecha = fecha + 3
    print(fecha)
    # if not validaciones.validar_si_no(data_cell):
    sheet.cell(row=r, column=column).value = fecha
      # file.write("Row: {} failed".format(r) + os.linesep)
    
 # Cerrar archivo de texto y excel
  workbook.save(path)
  file.close()
  print("data success")

validar_numero(workbook, sheet, 15, 1591, './log.txt')